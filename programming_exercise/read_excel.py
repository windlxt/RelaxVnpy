"""
作者：太乙真人
心境：行到水穷处 坐看云起时
日期：2024年06月04日
"""
import openpyxl
from PySide6.QtWidgets import QApplication, QTableWidget, QWidget, QVBoxLayout, \
    QHeaderView, QTableWidgetItem


class ExcelTableWidget(QWidget):
    def __init__(self, excel_file_path, parent=None):
        super().__init__(parent)
        self.tableWidget = QTableWidget()
        self.load_excel(excel_file_path)
        layout = QVBoxLayout()
        layout.addWidget(self.tableWidget)
        self.setLayout(layout)

    def load_excel(self, excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active

        # 设置QTableWidget的行和列数
        self.tableWidget.setRowCount(worksheet.max_row)
        self.tableWidget.setColumnCount(worksheet.max_column)

        # 设置列头
        self.tableWidget.setHorizontalHeaderLabels([col.value for col in worksheet[1]])
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        for row in range(1, worksheet.max_row + 1):
            for col in range(worksheet.max_column):
                # 读取单元格数据
                cell = worksheet.cell(row=row, column=col + 1)
                item_value = cell.value if cell.value is not None else ''
                # 在QTableWidget中设置数据
                self.tableWidget.setItem(row - 1, col, QTableWidgetItem(str(item_value)))


if __name__ == "__main__":
    app = QApplication([])
    excel_table_widget = ExcelTableWidget('/home/lxt/stock_plan/stock.xlsx')  # 替换为你的Excel文件路径
    excel_table_widget.show()
    app.exec()
