"""
作者：太乙真人
心境：行到水穷处 坐看云起时
日期：2024年05月19日
"""
import csv
from datetime import datetime
import platform
import openpyxl

from enum import Enum
from typing import Any, Dict, List
from copy import copy

import pandas as pd
from tzlocal import get_localzone_name

import importlib_metadata
from PySide6.QtWidgets import QWidget, QLabel, QTableWidget, QDialog, QHBoxLayout, \
    QVBoxLayout, QMenu, QTableWidgetItem, QHeaderView, QFileDialog, QLineEdit, QPushButton, \
    QFormLayout, QScrollArea, QMessageBox, QAbstractItemView
from PySide6.QtGui import QIcon, QColor, QAction, QContextMenuEvent, QCursor, QBrush
from PySide6.QtCore import Qt, Signal, QSettings, QByteArray, Slot

from vnpy.chart import StockChartWidget, CandleItem, VolumeItem
from ..constant import Direction, Exchange, Offset, OrderType
from ..engine import MainEngine, Event, EventEngine
from ..event import (
    EVENT_QUOTE,
    EVENT_TICK,
    EVENT_TRADE,
    EVENT_ORDER,
    EVENT_POSITION,
    EVENT_ACCOUNT,
    EVENT_LOG
)
from ..object import (
    OrderRequest,
    SubscribeRequest,
    CancelRequest,
    ContractData,
    PositionData,
    OrderData,
    QuoteData,
    TickData
)
from ..utility import load_json, save_json, get_digits, ZoneInfo, printr, printb, printg, printy
from ..setting import SETTING_FILENAME, SETTINGS
from ..locale import _

from vnpy.trader.database import BaseDatabase, get_database

EVENT_STOCK = "eStock."


COLOR_RED = QColor("red")
COLOR_GREEN = QColor("green")
COLOR_BID = QColor(255, 174, 201)
COLOR_ASK = QColor(160, 255, 160)
COLOR_BLACK = QColor("black")


class BaseCell(QTableWidgetItem):
    """
    General cell used in tablewidgets.
    """

    def __init__(self, content: Any, data: Any) -> None:
        """"""
        super().__init__()
        self.setTextAlignment(Qt.AlignCenter)
        self.set_content(content, data)

    def set_content(self, content: Any, data: Any) -> None:
        """
        Set text content.
        """
        self.setText(str(content))
        self._data = data

    def get_data(self) -> Any:
        """
        Get data object.
        """
        return self._data


class EnumCell(BaseCell):
    """
    Cell used for showing enum data.
    """

    def __init__(self, content: str, data: Any) -> None:
        """"""
        super().__init__(content, data)

    def set_content(self, content: Any, data: Any) -> None:
        """
        Set text using enum.constant.value.
        """
        if content:
            super().set_content(content.value, data)


class TimeCell(BaseCell):
    """
    Cell used for showing time string from datetime object.
    """

    local_tz = ZoneInfo(get_localzone_name())

    def __init__(self, content: Any, data: Any) -> None:
        """"""
        super().__init__(content, data)

    def set_content(self, content: Any, data: Any) -> None:
        """"""
        if content is None:
            return

        content: datetime = content.astimezone(self.local_tz)
        timestamp: str = content.strftime("%H:%M:%S")

        millisecond: int = int(content.microsecond / 1000)
        if millisecond:
            timestamp = f"{timestamp}.{millisecond}"
        else:
            timestamp = f"{timestamp}.000"

        self.setText(timestamp)
        self._data = data


class DateCell(BaseCell):
    """
    Cell used for showing date string from datetime object.
    """

    def __init__(self, content: Any, data: Any) -> None:
        """"""
        super().__init__(content, data)

    def set_content(self, content: Any, data: Any) -> None:
        """"""
        if content is None:
            return

        self.setText(content.strftime("%Y-%m-%d"))
        self._data = data


class MsgCell(BaseCell):
    """
    Cell used for showing msg data.
    """

    def __init__(self, content: str, data: Any) -> None:
        """"""
        super().__init__(content, data)
        self.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)


class HoldingDaysCell(BaseCell):
    """
    Cell used for showing date string from datetime object.
    """

    def __init__(self, content: Any, data: Any) -> None:
        """"""
        super().__init__(content, data)

    def set_content(self, content: Any, data: Any) -> None:
        """"""
        if content is None:
            return

        if int(content) > 15:
            self.setForeground(QBrush(COLOR_RED))

        self.setText(str(content))
        self._data = data


class ProfitLossCell(BaseCell):
    """
    Cell used for showing date string from datetime object.
    """

    def __init__(self, content: Any, data: Any) -> None:
        """"""
        super().__init__(content, data)

    def set_content(self, content: Any, data: Any) -> None:
        """"""
        if content is None:
            return

        if float(content) > 0:
            self.setForeground(QBrush(COLOR_RED))
        elif float(content) < 0:
            self.setForeground(QBrush(COLOR_GREEN))

        self.setText(content)
        self._data = data


class BaseMonitor(QTableWidget):
    """
    Monitor data update.
    """

    event_type: str = ""
    data_key: str = ""
    sorting: bool = False
    headers: dict = {}

    signal: Signal = Signal(Event)

    def __init__(self, main_engine: MainEngine, event_engine: EventEngine) -> None:
        """"""
        super().__init__()

        self.main_engine: MainEngine = main_engine
        self.event_engine: EventEngine = event_engine
        self.cells: Dict[str, dict] = {}

        self.database: BaseDatabase = get_database()

        self.init_ui()
        self.load_setting()
        self.register_event()

        self.itemClicked.connect(self.process_item_clicked)

    def init_ui(self) -> None:
        """"""
        self.init_table()
        self.init_menu()

    def init_table(self) -> None:
        """
        Initialize table.
        """
        self.setColumnCount(len(self.headers))

        labels: list = [d["display"] for d in self.headers.values()]
        self.setHorizontalHeaderLabels(labels)

        self.verticalHeader().setVisible(False)
        self.setEditTriggers(QAbstractItemView.AnyKeyPressed)
        self.setAlternatingRowColors(True)
        self.setSortingEnabled(self.sorting)

        self.insertRow(0)
        for column in range(self.columnCount()):
            item = QTableWidgetItem(f"新数据 {column}")
            self.setItem(0, column, item)

    def init_menu(self) -> None:
        """
        Create right click menu.
        """
        self.menu: QMenu = QMenu(self)

        resize_action: QAction = QAction(_("调整列宽"), self)
        resize_action.triggered.connect(self.resize_columns)
        self.menu.addAction(resize_action)

        save_action: QAction = QAction(_("保存数据"), self)
        save_action.triggered.connect(self.save_csv)
        self.menu.addAction(save_action)

    def register_event(self) -> None:
        """
        Register event handler into event engine.
        """
        if self.event_type:
            self.signal.connect(self.process_event)
            self.event_engine.register(self.event_type, self.signal.emit)

    def process_event(self, event: Event) -> None:
        """
        Process new data from event and update into table.
        """
        print('enter process_event.')
        # Disable sorting to prevent unwanted error.
        if self.sorting:
            self.setSortingEnabled(False)

        # Update data into table.
        data = event.data

        if not self.data_key:
            self.insert_new_row(data)
        else:
            key: str = data.__getattribute__(self.data_key)

            if key in self.cells:
                self.update_old_row(data)
            else:
                self.insert_new_row(data)

        # Enable sorting
        if self.sorting:
            self.setSortingEnabled(True)

    def insert_new_row(self, data: Any) -> None:
        """
        Insert a new row at the top of table.
        """
        self.insertRow(0)

        row_cells: dict = {}
        for column, header in enumerate(self.headers.keys()):
            setting: dict = self.headers[header]

            content = data.__getattribute__(header)
            cell: QTableWidgetItem = setting["cell"](content, data)
            self.setItem(0, column, cell)

            if setting["update"]:
                row_cells[header] = cell

        if self.data_key:
            key: str = data.__getattribute__(self.data_key)
            self.cells[key] = row_cells

    def update_old_row(self, data: Any) -> None:
        """
        Update an old row in table.
        """
        key: str = data.__getattribute__(self.data_key)
        row_cells = self.cells[key]

        for header, cell in row_cells.items():
            content = data.__getattribute__(header)
            cell.set_content(content, data)

    def resize_columns(self) -> None:
        """
        Resize all columns according to contents.
        """
        self.horizontalHeader().resizeSections(QHeaderView.ResizeToContents)

    def save_csv(self) -> None:
        """
        Save table data into a csv file
        """
        path, __ = QFileDialog.getSaveFileName(
            self, _("保存数据"), "", "CSV(*.csv)")

        if not path:
            return

        with open(path, "w") as f:
            writer = csv.writer(f, lineterminator="\n")

            headers: list = [d["display"] for d in self.headers.values()]
            writer.writerow(headers)

            for row in range(self.rowCount()):
                if self.isRowHidden(row):
                    continue

                row_data: list = []
                for column in range(self.columnCount()):
                    item: QTableWidgetItem = self.item(row, column)
                    if item:
                        row_data.append(str(item.text()))
                    else:
                        row_data.append("")
                writer.writerow(row_data)

    def contextMenuEvent(self, event: QContextMenuEvent) -> None:
        """
        Show menu with right click.
        """
        self.menu.popup(QCursor.pos())

    def save_setting(self) -> None:
        """"""
        settings: QSettings = QSettings(self.__class__.__name__, "custom")
        settings.setValue("column_state", self.horizontalHeader().saveState())

    def load_setting(self) -> None:
        """"""
        settings: QSettings = QSettings(self.__class__.__name__, "custom")
        column_state = settings.value("column_state")

        if isinstance(column_state, QByteArray):
            self.horizontalHeader().restoreState(column_state)
            self.horizontalHeader().setSortIndicator(-1, Qt.AscendingOrder)

    def process_item_clicked(self, item):
        pass


class StockHoldingPool(BaseMonitor):
    """持有股票池窗口，一日一更新"""
    event_type: str = EVENT_STOCK
    data_key: str = "code"
    sorting: bool = True

    headers: dict = {
        "date_buy": {"display": _("买入日期"), "cell": DateCell, "update": False},
        "code": {"display": _("代码"), "cell": BaseCell, "update": False},
        "exchange": {"display": _("交易所"), "cell": BaseCell, "update": False},
        "stock_name": {"display": _("股票名称"), "cell": BaseCell, "update": False},
        "price_cost": {"display": _("成本价"), "cell": BaseCell, "update": True},
        "volume_holding": {"display": _("持有量"), "cell": BaseCell, "update": True},
        "logic_buy": {"display": _("买入逻辑"), "cell": BaseCell, "update": False},
        "days_holding": {"display": _("持股天数"), "cell": HoldingDaysCell, "update": True},
        "price_current": {"display": _("现价"), "cell": BaseCell, "update": True},
        "cost": {"display": _("成本"), "cell": BaseCell, "update": True},
        "market_value": {"display": _("市价"), "cell": BaseCell, "update": True},
        "profit_loss": {"display": _("盈亏"), "cell": ProfitLossCell, "update": True}
        }

    def __init__(self, main_engine, event_engine, chart):
        self.chart = chart
        super().__init__(main_engine, event_engine)

    def init_table(self) -> None:
        self.setColumnCount(len(self.headers))

        labels: list = [d["display"] for d in self.headers.values()]
        self.setHorizontalHeaderLabels(labels)

        self.verticalHeader().setVisible(False)
        self.setEditTriggers(self.NoEditTriggers)
        self.setAlternatingRowColors(True)
        self.setSortingEnabled(self.sorting)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)  # 设置选择行为

        # 读取数据
        workbook = openpyxl.load_workbook('/home/lxt/stock_plan/stock.xlsx')
        worksheet = workbook.active
        # 转换成 DataFrame，用于计算
        self.df = pd.DataFrame(worksheet)

        # 设置QTableWidget的行和列数
        self.setRowCount(worksheet.max_row-1)
        self.setColumnCount(worksheet.max_column)

        # 设置列头
        # self.setHorizontalHeaderLabels([col.value for col in worksheet[1]])
        # self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        for row in range(2, worksheet.max_row + 1):
            for col in range(worksheet.max_column):
                # 读取单元格数据
                cell = worksheet.cell(row=row, column=col + 1)
                item_value = cell.value if cell.value is not None else ''
                # 格式化日期
                if col == 0:
                    setting: dict = self.headers['date_buy']
                    self.setItem(row - 2, col, setting["cell"](item_value, item_value))
                # 计算持股天数
                elif col == 7:
                    setting: dict = self.headers['days_holding']
                    day_plus = datetime.now() - self.df.iloc[row-1, 0].value
                    self.setItem(row - 2, col, setting["cell"](day_plus.days, day_plus.days))
                # 计算成本
                elif col == worksheet.max_column-3:
                    setting: dict = self.headers['cost']
                    cost = float(self.df.iloc[row-1, 4].value) * float(self.df.iloc[row-1, 5].value)
                    self.setItem(row - 2, col, setting["cell"]('{:.2f}'.format(cost), cost))
                # 计算市价
                elif col == worksheet.max_column - 2:
                    setting: dict = self.headers['market_value']
                    market_value = float(self.df.iloc[row - 1, 5].value) * float(self.df.iloc[row - 1, 8].value)
                    self.setItem(row - 2, col, setting["cell"]('{:.2f}'.format(market_value), market_value))
                # 计算盈亏
                elif col == worksheet.max_column - 1:
                    setting: dict = self.headers['profit_loss']
                    profit_loss = market_value - cost
                    self.setItem(row - 2, col, setting["cell"]('{:.2f}'.format(profit_loss), profit_loss))
                else:
                    # 在QTableWidget中设置数据
                    match col:
                        case 1:
                            colume = 'code'
                        case 2:
                            colume = 'exchange'
                        case 3:
                            colume = 'stock_name'
                        case 4:
                            colume = 'price_cost'
                        case 5:
                            colume = 'volume_holding'
                        case 6:
                            colume = 'logic_buy'
                        case 8:
                            colume = 'price_current'
                    setting: dict = self.headers[colume]

                    self.setItem(row - 2, col, setting['cell'](item_value, item_value))

        self.resize_columns()

        self.process_item_clicked(0)

    @Slot()
    def process_item_clicked(self, item):

        if item == 0:
            code = str(self.df.iloc[1, 2].value) + '.' + str(self.df.iloc[1, 1].value)
            print(code)
        else:
            code = str(self.df.iloc[item.row()+1, 2].value) + '.' + str(self.df.iloc[item.row()+1, 1].value)
        self.startdate: str = SETTINGS["database.startdate"]

        history = self.database.load_bar_data(code, '', '', datetime.strptime(self.startdate, '%Y-%m-%d'), datetime.now())

        try:
            print(history[0])
        except IndexError:
            self.chart.clear()
            printr('数据库没有存储该股票数据！')

        self.chart.update_history(history)


class StockPrepareSellPool(BaseMonitor):
    """准备卖出股票池窗口，一日一更新"""
    event_type: str = ""
    data_key: str = "code"
    sorting: bool = True

    headers: dict = {
        "today": {"display": _("卖出日期"), "cell": TimeCell, "update": True},
        "code": {"display": _("代码"), "cell": BaseCell, "update": False},
        "stock_name": {"display": _("股票名称"), "cell": EnumCell, "update": False},
        "price_buy": {"display": _("成本"), "cell": BaseCell, "update": True},
        "volume_holding": {"display": _("持有量"), "cell": BaseCell, "update": True},
        "days_holding": {"display": _("持股天数"), "cell": BaseCell, "update": True},
        "logic_sell": {"display": _("卖出逻辑"), "cell": BaseCell, "update": True},
        "price_sell": {"display": _("拟卖出价格"), "cell": BaseCell, "update": True},
        "volume_sell": {"display": _("拟卖出量"), "cell": BaseCell, "update": True},
    }


class StockPrepareBuyPool(BaseMonitor):
    """准备买入股票池窗口，一日一更新"""
    event_type: str = ""
    data_key: str = "code"
    sorting: bool = True

    headers: dict = {
        "today": {"display": _("买入日期"), "cell": TimeCell, "update": True},
        "code": {"display": _("代码"), "cell": BaseCell, "update": False},
        "stock_name": {"display": _("股票名称"), "cell": EnumCell, "update": False},
        "price_buy": {"display": _("拟买入价格"), "cell": BaseCell, "update": True},
        "volume_buy": {"display": _("拟买入量"), "cell": BaseCell, "update": True},
        "logic_buy": {"display": _("买入逻辑"), "cell": BaseCell, "update": True},
    }

    def update_with_cell(self):
        """更新窗口内容"""
        print('更新 StockPool 窗口内容。')


class StockWatchPool(BaseMonitor):
    """观察股票池窗口，一日一更新"""
    event_type: str = ""
    data_key: str = "code"
    sorting: bool = True

    headers: dict = {
        "code": {"display": _("代码"), "cell": BaseCell, "update": False},
        "stock_name": {"display": _("股票名称"), "cell": EnumCell, "update": False},
    }


class StockAllPool(QWidget):
    """所有股票池窗口，策略回测和筛查"""
    def __init__(self, main_engine, event_engine):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.label = QLabel('所有股票池窗口，策略回测和筛查')

        self.vlayout = QVBoxLayout()
        self.vlayout.addWidget(self.label)
        self.setLayout(self.vlayout)


class StockSelfSelectedPool(QWidget):
    """自选股票池窗口，策略回测和筛查"""
    def __init__(self, main_engine, event_engine):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.label = QLabel('自选股票池窗口，策略回测和筛查')

        self.vlayout = QVBoxLayout()
        self.vlayout.addWidget(self.label)
        self.setLayout(self.vlayout)


class StockKLineChart(QWidget):
    """股票K线图"""
    def __init__(self, main_engine, event_engine):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        # Create chart widget
        self.chart: StockChartWidget = StockChartWidget()
        self.chart.add_plot("candle", hide_x_axis=True)
        self.chart.add_plot("volume", maximum_height=200)
        self.chart.add_item(CandleItem, "candle", "candle")
        self.chart.add_item(VolumeItem, "volume", "volume")
        self.chart.add_cursor()

        vbox: QVBoxLayout = QVBoxLayout()
        vbox.addWidget(self.chart)
        self.setLayout(vbox)


        # # Create help widget
        # text1: str = "红色虚线 —— 盈利交易"
        # label1: QLabel = QLabel(text1)
        # label1.setStyleSheet("color:red")
        #
        # text2: str = "绿色虚线 —— 亏损交易"
        # label2: QLabel = QLabel(text2)
        # label2.setStyleSheet("color:#00FF00")
        #
        # text3: str = "黄色向上箭头 —— 买入开仓 Buy"
        # label3: QLabel = QLabel(text3)
        # label3.setStyleSheet("color:yellow")
        #
        # text4: str = "黄色向下箭头 —— 卖出平仓 Sell"
        # label4: QLabel = QLabel(text4)
        # label4.setStyleSheet("color:yellow")
        #
        # text5: str = "紫红向下箭头 —— 卖出开仓 Short"
        # label5: QLabel = QLabel(text5)
        # label5.setStyleSheet("color:magenta")
        #
        # text6: str = "紫红向上箭头 —— 买入平仓 Cover"
        # label6: QLabel = QLabel(text6)
        # label6.setStyleSheet("color:magenta")
        #
        # hbox1: QHBoxLayout = QHBoxLayout()
        # hbox1.addStretch()
        # hbox1.addWidget(label1)
        # hbox1.addStretch()
        # hbox1.addWidget(label2)
        # hbox1.addStretch()
        #
        # hbox2: QHBoxLayout = QHBoxLayout()
        # hbox2.addStretch()
        # hbox2.addWidget(label3)
        # hbox2.addStretch()
        # hbox2.addWidget(label4)
        # hbox2.addStretch()
        #
        # hbox3: QHBoxLayout = QHBoxLayout()
        # hbox3.addStretch()
        # hbox3.addWidget(label5)
        # hbox3.addStretch()
        # hbox3.addWidget(label6)
        # hbox3.addStretch()
        #
        # # Set layout
        # vbox: QVBoxLayout = QVBoxLayout()
        # vbox.addWidget(self.chart)
        # vbox.addLayout(hbox1)
        # vbox.addLayout(hbox2)
        # vbox.addLayout(hbox3)
        # self.setLayout(vbox)


class ContractManager(QWidget):
    """
    Query contract data available to trade in system.
    """

    headers: Dict[str, str] = {
        "vt_symbol": _("本地代码"),
        "symbol": _("代码"),
        "exchange": _("交易所"),
        "name": _("名称"),
        "product": _("合约分类"),
        "size": _("合约乘数"),
        "pricetick": _("价格跳动"),
        "min_volume": _("最小委托量"),
        "option_portfolio": _("期权产品"),
        "option_expiry": _("期权到期日"),
        "option_strike": _("期权行权价"),
        "option_type": _("期权类型"),
        "gateway_name": _("交易接口"),
    }

    def __init__(self, main_engine: MainEngine, event_engine: EventEngine) -> None:
        super().__init__()

        self.main_engine: MainEngine = main_engine
        self.event_engine: EventEngine = event_engine

        self.init_ui()

    def init_ui(self) -> None:
        """"""
        self.setWindowTitle(_("合约查询"))
        self.resize(1000, 600)

        self.filter_line: QLineEdit = QLineEdit()
        self.filter_line.setPlaceholderText(_("输入合约代码或者交易所，留空则查询所有合约"))

        self.button_show: QPushButton = QPushButton(_("查询"))
        self.button_show.clicked.connect(self.show_contracts)

        labels: list = []
        for name, display in self.headers.items():
            label: str = f"{display}\n{name}"
            labels.append(label)

        self.contract_table: QTableWidget = QTableWidget()
        self.contract_table.setColumnCount(len(self.headers))
        self.contract_table.setHorizontalHeaderLabels(labels)
        self.contract_table.verticalHeader().setVisible(False)
        self.contract_table.setEditTriggers(self.contract_table.NoEditTriggers)
        self.contract_table.setAlternatingRowColors(True)

        hbox: QHBoxLayout = QHBoxLayout()
        hbox.addWidget(self.filter_line)
        hbox.addWidget(self.button_show)

        vbox: QVBoxLayout = QVBoxLayout()
        vbox.addLayout(hbox)
        vbox.addWidget(self.contract_table)

        self.setLayout(vbox)

    def show_contracts(self) -> None:
        """
        Show contracts by symbol
        """
        flt: str = str(self.filter_line.text())

        all_contracts: List[ContractData] = self.main_engine.get_all_contracts()
        if flt:
            contracts: List[ContractData] = [
                contract for contract in all_contracts if flt in contract.vt_symbol
            ]
        else:
            contracts: List[ContractData] = all_contracts

        self.contract_table.clearContents()
        self.contract_table.setRowCount(len(contracts))

        for row, contract in enumerate(contracts):
            for column, name in enumerate(self.headers.keys()):
                value: object = getattr(contract, name)

                if value in {None, 0, 0.0}:
                    value = ""

                if isinstance(value, Enum):
                    cell: EnumCell = EnumCell(value, contract)
                elif isinstance(value, datetime):
                    cell: DateCell = DateCell(value, contract)
                else:
                    cell: BaseCell = BaseCell(value, contract)
                self.contract_table.setItem(row, column, cell)

        self.contract_table.resizeColumnsToContents()


class GlobalDialog(QDialog):
    """
    Start connection of a certain gateway.
    """

    def __init__(self) -> None:
        """"""
        super().__init__()

        self.widgets: Dict[str, Any] = {}

        self.init_ui()

    def init_ui(self) -> None:
        """"""
        self.setWindowTitle(_("全局配置"))
        self.setMinimumWidth(800)

        settings: dict = copy(SETTINGS)
        settings.update(load_json(SETTING_FILENAME))

        # Initialize line edits and form layout based on setting.
        form: QFormLayout = QFormLayout()

        for field_name, field_value in settings.items():
            field_type: type = type(field_value)
            widget: QLineEdit = QLineEdit(str(field_value))

            form.addRow(f"{field_name} <{field_type.__name__}>", widget)
            self.widgets[field_name] = (widget, field_type)

        button: QPushButton = QPushButton(_("确定"))
        button.clicked.connect(self.update_setting)
        form.addRow(button)

        scroll_widget: QWidget = QWidget()
        scroll_widget.setLayout(form)

        scroll_area: QScrollArea = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(scroll_widget)

        vbox: QVBoxLayout = QVBoxLayout()
        vbox.addWidget(scroll_area)
        self.setLayout(vbox)

    def update_setting(self) -> None:
        """
        Get setting value from line edits and update global setting file.
        """
        settings: dict = {}
        for field_name, tp in self.widgets.items():
            widget, field_type = tp
            value_text: str = widget.text()

            if field_type == bool:
                if value_text == "True":
                    field_value: bool = True
                else:
                    field_value: bool = False
            else:
                field_value = field_type(value_text)

            settings[field_name] = field_value

        QMessageBox.information(
            self,
            "注意",
            "全局配置的修改需要重启后才会生效！",
            QMessageBox.Ok
        )

        save_json(SETTING_FILENAME, settings)
        self.accept()


class AboutDialog(QDialog):
    """
    Information about the trading platform.
    """

    def __init__(self, main_engine: MainEngine, event_engine: EventEngine) -> None:
        """"""
        super().__init__()

        self.main_engine: MainEngine = main_engine
        self.event_engine: EventEngine = event_engine

        self.init_ui()

    def init_ui(self) -> None:
        """"""
        self.setWindowTitle(_("关于VeighNa Trader"))

        from ... import __version__ as vnpy_version

        text: str = f"""
            By Traders, For Traders.

            Created by VeighNa Technology


            License：MIT
            Website：www.vnpy.com
            Github：www.github.com/vnpy/vnpy


            VeighNa - {vnpy_version}
            Python - {platform.python_version()}
            PySide6 - {importlib_metadata.version("pyside6")}
            NumPy - {importlib_metadata.version("numpy")}
            pandas - {importlib_metadata.version("pandas")}
            """

        label: QLabel = QLabel()
        label.setText(text)
        label.setMinimumWidth(500)

        vbox: QVBoxLayout = QVBoxLayout()
        vbox.addWidget(label)
        self.setLayout(vbox)


